import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
import os

# Function to examine workbook properties
def examine_workbook(filepath):
    print(f"\nExamining: {filepath}")
    try:
        # Load without data_only to preserve formulas
        wb = openpyxl.load_workbook(filepath, data_only=False)
        
        # Get basic information
        print(f"Number of worksheets: {len(wb.worksheets)}")
        print(f"Active sheet: {wb.active.title}")
        
        # Loop through each worksheet
        for sheet in wb.worksheets:
            print(f"\nWorksheet: {sheet.title}")
            print(f"  Dimensions: {sheet.dimensions}")
            print(f"  Number of rows: {sheet.max_row}")
            print(f"  Number of columns: {sheet.max_column}")
            
            # Check for drawings (shapes, charts, images)
            if sheet._charts or sheet._images:
                print(f"  Contains {len(sheet._charts)} charts and {len(sheet._images)} images")
            
            # Check for merged cells
            if sheet.merged_cells:
                print(f"  Contains {len(sheet.merged_cells.ranges)} merged cell ranges")
                
            # Check for conditional formatting
            if sheet.conditional_formatting:
                print(f"  Contains {len(sheet.conditional_formatting)} conditional formatting rules")
                
            # Look for formulas in first few rows
            formula_count = 0
            for row in range(1, min(20, sheet.max_row + 1)):
                for col in range(1, min(10, sheet.max_column + 1)):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
            
            if formula_count > 0:
                print(f"  Contains at least {formula_count} formulas in the first few rows")
    
    except Exception as e:
        print(f"Error examining {filepath}: {e}")

# Check both the original and the processed file
original_file = '4_1.xlsx'
processed_file = 'tmp66qmxn_6.xlsx'

if os.path.exists(original_file):
    examine_workbook(original_file)
else:
    print(f"Original file {original_file} not found!")

if os.path.exists(processed_file):
    examine_workbook(processed_file)
else:
    print(f"Processed file {processed_file} not found!")

# Check if there are specific differences in cell content in the first few rows
print("\nComparing cell content in the first 10 rows:")
try:
    wb_orig = openpyxl.load_workbook(original_file, data_only=True)
    wb_proc = openpyxl.load_workbook(processed_file, data_only=True)
    
    sheet_orig = wb_orig.active
    sheet_proc = wb_proc.active
    
    max_row = min(20, sheet_orig.max_row, sheet_proc.max_row)
    max_col = min(10, sheet_orig.max_column, sheet_proc.max_column)
    
    diff_count = 0
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            val_orig = sheet_orig.cell(row=row, column=col).value
            val_proc = sheet_proc.cell(row=row, column=col).value
            
            if val_orig != val_proc:
                print(f"  Difference at row {row}, col {col}:")
                print(f"    Original: {val_orig}")
                print(f"    Processed: {val_proc}")
                diff_count += 1
                if diff_count >= 5:  # Limit to first 5 differences
                    print("  (showing only first 5 differences)")
                    break
        if diff_count >= 5:
            break
            
except Exception as e:
    print(f"Error comparing files: {e}") 