"""
Обработчик для файлов формата 4_1.xlsx - упрощенная версия
"""

import pandas as pd
import os
import re
import openpyxl
import shutil
import time
from copy import copy
from main import Processor, group_similar
from src.morphology import normalize_term
from src.okpd_fetch import fetch_okpd2_batch
from openpyxl.utils import get_column_letter
from .base_processor import BaseProcessor

class FullFormatProcessor(BaseProcessor):
    """
    Процессор для формата 4_1.
    
    Просто извлекает данные из колонки 'Наименование' и записывает коды в 'Код ОКП/ОКПД2',
    при этом пропускает служебные строки.
    """
    
    # Количество строк сверху, которые нужно пропустить (заголовки таблицы)
    # Определяется как атрибут класса, но может быть переопределен для экземпляра
    _DEFAULT_HEADER_ROWS = 5
    
    # Шаблоны для строк, которые нужно пропустить
    SKIP_PATTERNS = [
        # Полные фразы
        r'ВСЕГО\s+по\s+разделу(\s+\d+)?',
        r'ИТОГО\s+по\s+разделу(\s+\d+)?',
        r'ВСЕГО\s+\d+',
        r'ИТОГО\s+\d+',
        r'Сырье\s+и\s+основные\s+материалы',
        r'Вспомогательные\s+материалы',
        r'Возвратные\s+отходы',
        r'Приобретение\s+комплектующих\s+изделий',
        r'Покупные\s+комплектующие\s+изделия',
        r'Возвратные\s+отходы\s+\(вычитаются\)',
        # Заголовки разделов
        r'Раздел\s+\d+',
        r'^\s*№\s*п/п\s*$',
        r'\bНаименование\s+показателя\b',
        # Отдельные ключевые слова с учетом границ слов
        r'\bИТОГО\b',
        r'\bВСЕГО\b',
        r'\bСырье\b', 
        r'\bВспомогательные(\s+материалы)?\b',
        r'\bВозвратные(\s+отходы)?\b',
        r'\bПриобретение\b',
        r'\bПокупные\b',
        r'\bОтходы\b',
        r'\bМатериалы\b',
        r'\bКомплектующие\b',
        r'\bПолуфабрикаты\b',
        r'\bИзделия\b',
        # Дополнительные служебные тексты
        r'Код\s+ОКП',
        r'Код\s+ОКПД',
        r'Единица\s+измерения',
        r'\bТС\b',
        r'\bШт\b$'
    ]
    
    # Компилируем регулярные выражения для быстрой проверки
    SKIP_PATTERNS_COMPILED = [re.compile(pattern, re.IGNORECASE) for pattern in SKIP_PATTERNS]
    
    def __init__(self, input_file=None, checkpoint_name="checkpoint.xlsx", save_interval=10, progress=None):
        super().__init__(input_file, checkpoint_name, save_interval, progress)
        # Инициализируем напрямую без использования свойства
        self._num_header_rows = self._DEFAULT_HEADER_ROWS
        self.skipped_rows = 0
        self.skipped_service_rows = 0
        self.processed_items = 0
        
        # Создаем резервную копию входного файла сразу
        if input_file and input_file.name:
            self._create_backup_file()
            
        # Коды ОКПД для обновления в файле
        self.results_to_update = {}
        
        # Данные о файле
        self.workbook = None
        self.sheet_name = None
        self.name_column_index = None
        self.code_column_index = None
    
    # Свойство для доступа к атрибуту _num_header_rows экземпляра
    @property
    def NUM_HEADER_ROWS(self):
        return self._num_header_rows
        
    @NUM_HEADER_ROWS.setter
    def NUM_HEADER_ROWS(self, value):
        self._num_header_rows = int(value)
        self.logger.info(f"Установлено пропуск {self._num_header_rows} строк заголовка")
        
    def _create_backup_file(self):
        """Создает резервную копию исходного файла"""
        try:
            original_path = self.input_path
            if not original_path or not os.path.exists(original_path):
                return
                
            # Формируем имя для бэкапа
            base_name, ext = os.path.splitext(original_path)
            backup_path = f"{base_name}_original{ext}"
            
            # Создаем копию только если её ещё нет
            if not os.path.exists(backup_path):
                shutil.copy2(original_path, backup_path)
                self.logger.info(f"Создана резервная копия исходного файла: {backup_path}")
                return backup_path
        except Exception as e:
            self.logger.warning(f"Не удалось создать резервную копию: {e}")
        return None
        
    def _find_columns_in_excel(self):
        """
        Находит колонки с наименованием и кодами ОКПД в файле Excel
        непосредственно используя openpyxl
        """
        if not self.input_path or not os.path.exists(self.input_path):
            self.logger.error("Не указан путь к входному файлу")
            return False
            
        try:
            # Открываем Excel файл
            if not self.workbook:
                self.workbook = openpyxl.load_workbook(self.input_path)
            self.logger.info(f"Excel файл открыт: {self.input_path}")
            
            # Определяем нужный лист на основе self.sheet_name
            if self.sheet_name and self.sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[self.sheet_name]
            else:
                # Используем активный лист, если sheet_name не задан или не найден
                sheet = self.workbook.active
                self.sheet_name = sheet.title
            
            self.logger.info(f"Анализируем лист: {sheet.title}")
            
            # Ищем в первых нескольких строках заголовки колонок
            name_column = None
            code_column = None
            
            # Проверяем первые 15 строк (обычно заголовки там)
            max_check_row = min(15, sheet.max_row)
            
            for row in range(1, max_check_row + 1):
                for col in range(1, min(20, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    
                    if not cell_value:
                        continue
                        
                    cell_text = str(cell_value).strip().lower()
                    
                    # Ищем колонку с наименованием
                    if cell_text == "наименование" or "наименов" in cell_text:
                        name_column = col
                        self.logger.info(f"Найдена колонка 'Наименование': строка {row}, колонка {col}")
                    
                    # Ищем колонку с кодом ОКПД
                    if "код" in cell_text and ("окп" in cell_text or "окпд" in cell_text):
                        code_column = col
                        self.logger.info(f"Найдена колонка 'Код ОКП/ОКПД2': строка {row}, колонка {col}")
                    
                    # Если нашли обе колонки, завершаем поиск
                    if name_column and code_column:
                        self.name_column_index = name_column
                        self.code_column_index = code_column
                        return True
            
            # Если не нашли колонку кода, но нашли наименование
            if name_column and not code_column:
                self.name_column_index = name_column
                # Ищем первую пустую колонку после наименования для кодов ОКПД
                for col in range(name_column + 1, min(name_column + 5, sheet.max_column + 1)):
                    is_empty = True
                    for row in range(self._num_header_rows + 1, min(self._num_header_rows + 10, sheet.max_row + 1)):
                        if sheet.cell(row=row, column=col).value:
                            is_empty = False
                            break
                    
                    if is_empty:
                        code_column = col
                        self.code_column_index = col
                        self.logger.info(f"Не найдена колонка для кодов ОКПД, будем использовать колонку {col}")
                        
                        # Добавляем заголовок в ячейку для кодов ОКПД
                        header_row = self._num_header_rows
                        for row in range(1, header_row + 1):
                            if sheet.cell(row=row, column=name_column).value:
                                sheet.cell(row=row, column=col).value = "Код ОКП/ОКПД2"
                                self.logger.info(f"Добавлен заголовок 'Код ОКП/ОКПД2' в строку {row}, колонку {col}")
                                break
                        return True
            
            # Не нашли нужные колонки
            self.logger.error(f"Не удалось найти колонки 'Наименование' и 'Код ОКП/ОКПД2' на листе {sheet.title}")
            return False
            
        except Exception as e:
            self.logger.exception(f"Ошибка при анализе Excel файла: {e}")
            return False
    
    def _process_file(self):
        """Обработка файла формата 4_1"""
        try:
            self.logger.info(f"Начало обработки файла формата 4_1: {self.input_path}")
            self.logger.info(f"Пропускаем первые {self._num_header_rows} строк заголовка")
            
            # Открываем Excel файл для получения информации о листах
            if not os.path.exists(self.input_path):
                self.logger.error(f"Файл не найден: {self.input_path}")
                return False
                
            try:
                # Открываем файл с помощью openpyxl для получения списка листов
                workbook = openpyxl.load_workbook(self.input_path)
                sheet_names = workbook.sheetnames
                sheet_count = len(sheet_names)
                self.logger.info(f"Файл содержит {sheet_count} листов: {', '.join(sheet_names)}")
                
                # Сохраняем workbook для дальнейшего использования
                self.workbook = workbook
                
                # Новый подход: собираем все элементы со всех листов сначала
                all_items = []  # Список всех элементов [{'text': '...', 'sheet': '...', 'row': N}]
                all_unique_items = set()  # Множество уникальных текстов элементов
                
                # 1. Собираем элементы со всех листов
                for sheet_idx, sheet_name in enumerate(sheet_names, start=1):
                    self.logger.info(f"Сканирование листа {sheet_idx}/{sheet_count}: '{sheet_name}'")
                    
                    # Устанавливаем текущий лист
                    self.sheet_name = sheet_name
                    
                    # Очищаем предыдущие результаты
                    self.name_column_index = None
                    self.code_column_index = None
                    
                    # Находим колонки в текущем листе
                    success = self._find_columns_in_excel()
                    if not success:
                        self.logger.warning(f"Невозможно обработать лист '{sheet_name}', пропускаем")
                        continue
                    
                    # Читаем данные из текущего листа
                    try:
                        # Читаем данные из конкретного листа
                        self.df = pd.read_excel(self.input_path, sheet_name=sheet_name)
                        self.logger.info(f"Лист '{sheet_name}' прочитан для анализа, обнаружено {self.df.shape[0]} строк, {self.df.shape[1]} столбцов")
                        
                        # Собираем элементы с текущего листа
                        sheet_items = self._collect_items_from_sheet()
                        
                        for item in sheet_items:
                            all_items.append({
                                'text': item[1],
                                'sheet': sheet_name,
                                'row': item[0]
                            })
                            all_unique_items.add(item[1])
                            
                        self.logger.info(f"Найдено {len(sheet_items)} элементов на листе '{sheet_name}'")
                        
                    except Exception as e:
                        self.logger.warning(f"Ошибка при чтении листа '{sheet_name}': {e}")
                        continue
                
                # Статистика собранных элементов
                self.logger.info(f"Всего найдено {len(all_items)} элементов на всех листах")
                self.logger.info(f"Уникальных элементов: {len(all_unique_items)}")
                
                if not all_unique_items:
                    self.logger.warning("Не найдено элементов для обработки")
                    return False
                
                # 2. Группируем похожие элементы
                unique_texts = list(all_unique_items)
                self.logger.info("Группируем похожие элементы...")
                
                # Инициализируем модель перед группировкой для экономии времени
                if not self.init_model():
                    self.logger.error("Ошибка инициализации модели")
                    return False
                
                # Группировка похожих элементов
                groups = group_similar(unique_texts)
                self.logger.info(f"Сгруппировано в {len(groups)} групп")
                
                # 3. Обрабатываем каждую группу и получаем коды ОКПД
                codes_by_item = {}  # Словарь {текст: код_ОКПД}
                
                # Безопасная работа с прогрессом
                if self.progress is not None:
                    self.progress(0, desc=f"Обработка групп...", total=len(groups))
                    progress_iter = self.progress.tqdm(groups, desc=f"Получение кодов ОКПД")
                else:
                    progress_iter = groups
                
                for idx, group in enumerate(progress_iter, start=1):
                    if self.stop_event.is_set():
                        self.logger.info("Обработка остановлена пользователем")
                        break
                        
                    if not group:
                        continue
                    
                    # Берем представителя группы
                    rep = group[0]
                    
                    # Дополнительная проверка перед обработкой
                    # Проверяем снова чтобы не пропустить служебные строки
                    is_service_line = False
                    for i, pattern in enumerate(self.SKIP_PATTERNS_COMPILED):
                        if pattern.search(rep):
                            self.logger.warning(f"Пропускаем служебную строку (повторная проверка): '{rep}' (соответствует шаблону {i+1}: {self.SKIP_PATTERNS[i]})")
                            is_service_line = True
                            break
                    
                    if is_service_line:
                        continue
                    
                    try:
                        # Нормализуем термин
                        normalized = normalize_term(rep)
                        self.logger.info(f"Обработка группы {idx}/{len(groups)}: {normalized}")
                        
                        # Получаем упрощенный термин
                        prompt = [
                            {"role": "system", "content": 'Ты помогаешь выбрать один код для военной компании, которая занимается производством и работает с различным металом, где производят Системы термостатирования и контроля температурно влажностного режима.'},
                            {"role": "user", "content": f"Перефразируй название товара, удалив все размеры и числовые параметры, преобразовав тип товара.\nЕсли встречаешь металические изделия, то прибавляй алюминевый. \n \
                            Если слово 'лист' -> 'профиль алюминевый', если слово 'круг' -> 'профиль алюминевый, если слово 'болт' или 'винт -> 'болты и винты', если слово 'гвоздь' -> 'гвоздь', если слово 'доска' или 'брусок' -> 'пиломатериалы', если слово 'жгут' -> 'жгуты синтетические', если слово 'бензин' -> 'бензин', если слово 'бензин' -> 'бензин'.\n \
                            Если встречаешь слово на английском языке - ничего не меняй. Напрмиер: если слово 'Isolontape 500 3005 VB D LM' -> 'Isolontape' \
                            Если встречаешь слово которого нет в примерах, ориентируйся и сделай на подобии. \
                            \nНазвание: {normalized}\nВыведи только товар:"}
                            ]
                        simplified = self.model.generate(prompt)['content']
                        self.logger.info(f"Упрощено до: {simplified}")
                        
                        # Запрашиваем коды ОКПД
                        okpd_data = fetch_okpd2_batch([simplified])
                        entries = okpd_data.get(simplified, [])
                        
                        # Выбираем подходящий код
                        code, name, comment = Processor._decide(self, entries, rep, simplified)
                        self.logger.info(f"Выбран код: {code} - {name}")
                        
                        # Сохраняем код для каждого элемента в группе
                        for item in group:
                            codes_by_item[item] = code
                            
                    except Exception as e:
                        self.logger.exception(f"Ошибка при обработке группы {idx}: {e}")
                
                # 4. Проставляем коды для всех вхождений элементов
                self.logger.info(f"Определены коды ОКПД для {len(codes_by_item)} уникальных элементов")
                
                # Создаем словарь для всех обновлений {sheet_name: {row: code}}
                updates_by_sheet = {}
                
                # Проходим по всем найденным элементам и составляем план обновлений
                for item in all_items:
                    item_text = item['text']
                    sheet_name = item['sheet']
                    row = item['row']
                    
                    if item_text in codes_by_item:
                        code = codes_by_item[item_text]
                        
                        # Инициализируем словарь для листа, если его еще нет
                        if sheet_name not in updates_by_sheet:
                            updates_by_sheet[sheet_name] = {}
                            
                        # Добавляем обновление для данной строки
                        updates_by_sheet[sheet_name][row] = {
                            'code': code,
                            'item': item_text
                        }
                
                # 5. Применяем обновления для каждого листа
                total_updated = 0
                
                for sheet_name, updates in updates_by_sheet.items():
                    self.logger.info(f"Обновление листа '{sheet_name}': {len(updates)} элементов")
                    
                    # Переключаемся на нужный лист
                    self.sheet_name = sheet_name
                    
                    # Явно получаем лист из рабочей книги
                    if sheet_name in self.workbook.sheetnames:
                        sheet = self.workbook[sheet_name]
                    else:
                        self.logger.warning(f"Лист '{sheet_name}' не найден в рабочей книге, пропускаем")
                        continue
                    
                    # Перед обновлением листа, находим колонки заново
                    success = self._find_columns_in_excel()
                    if not success:
                        self.logger.warning(f"Не удалось определить колонки на листе '{sheet_name}', пропускаем обновления")
                        continue
                    
                    # Проверяем колонки перед обновлением
                    if not self.name_column_index or not self.code_column_index:
                        self.logger.warning(f"Не удалось определить колонки на листе '{sheet_name}', пропускаем")
                        continue
                    
                    self.logger.info(f"Найдены колонки на листе '{sheet_name}': наименование({self.name_column_index}), код({self.code_column_index})")
                    
                    # Обновляем каждую ячейку
                    sheet_updated = 0
                    for row, data in updates.items():
                        try:
                            # Ячейка с кодом - добавляем +2 к индексу строки (+1 для нумерации Excel с 1, +1 для коррекции смещения)
                            code_cell = sheet.cell(row=int(row)+2, column=self.code_column_index)
                            old_value = code_cell.value
                            
                            # Для отладки: проверяем содержимое ячейки с наименованием
                            name_cell = sheet.cell(row=int(row)+2, column=self.name_column_index)
                            name_value = name_cell.value
                            if name_value != data['item']:
                                self.logger.info(f"Ожидаемый текст '{data['item']}', фактический '{name_value}' в строке {int(row)+2}")
                            
                            # Проверяем, содержит ли ячейка формулу
                            if old_value and isinstance(old_value, str) and old_value.startswith('='):
                                self.logger.warning(f"Ячейка ({sheet_name}:{row+2}, {self.code_column_index}) содержит формулу, пропускаем: {old_value}")
                                continue
                            
                            # Обновляем значение
                            code_cell.value = data['code']
                            total_updated += 1
                            sheet_updated += 1
                            
                            if total_updated % 100 == 0:
                                self.logger.info(f"Обновлено {total_updated} ячеек...")
                                
                            if sheet_updated <= 5:  # Логируем первые 5 обновлений для каждого листа
                                self.logger.info(f"Обновлена ячейка {sheet_name}:({int(row)+2}, {self.code_column_index}): '{old_value}' -> '{data['code']}' для '{name_value}'")
                                
                        except Exception as e:
                            self.logger.warning(f"Ошибка при обновлении ячейки на листе '{sheet_name}', строка {row+2}: {e}")
                    
                    self.logger.info(f"Обновлено {sheet_updated} ячеек на листе '{sheet_name}'")
                
                # 6. Сохраняем результаты
                try:
                    # Увеличиваем ширину колонки для кодов ОКПД
                    self._adjust_column_width()
                    
                    self.workbook.save(self.input_path)
                    self.logger.info(f"Файл успешно обновлен, проставлено {total_updated} кодов ОКПД")
                except Exception as e:
                    self.logger.error(f"Ошибка при сохранении файла: {e}")
                    return False
                
                # Копируем исходный файл в output_path для интерфейса
                try:
                    shutil.copy2(self.input_path, self.output_path)
                    self.logger.info(f"Копия результата сохранена в {self.output_path}")
                except Exception as e:
                    self.logger.warning(f"Не удалось создать копию результата: {e}")
                
                return True
                
            except Exception as e:
                self.logger.exception(f"Ошибка при обработке листов Excel: {e}")
                return False
            
        except Exception as e:
            self.logger.exception(f"Ошибка в FullFormatProcessor: {e}")
            return False
            
    def _collect_items_from_sheet(self):
        """
        Собирает элементы с текущего листа, пропуская служебные строки
        
        Returns:
            list: Список кортежей (индекс_строки, текст_элемента)
        """
        items_collected = []
        
        # Счетчики для анализа пропущенных строк
        headers_skipped = 0
        empty_skipped = 0
        patterns_skipped = 0
        total_rows = 0
        
        # Проверяем, определены ли колонки
        if self.name_column_index is None:
            self.logger.error(f"Не определен индекс колонки с наименованиями для листа {self.sheet_name}")
            return []
        
        # Определяем имя колонки с наименованиями
        try:
            if 'Наименование' in self.df.columns:
                item_column_name = 'Наименование'
            elif self.name_column_index and self.name_column_index - 1 < len(self.df.columns):
                # Преобразуем индекс колонки openpyxl (с 1) в индекс pandas (с 0)
                item_column_name = self.df.columns[self.name_column_index - 1]
            else:
                # Попробуем найти колонку по ключевым словам
                for col in self.df.columns:
                    if 'наименов' in str(col).lower():
                        item_column_name = col
                        self.logger.info(f"Найдена колонка c наименованиями по ключевому слову: {col}")
                        break
                else:
                    self.logger.error(f"Не удалось определить колонку с наименованиями на листе {self.sheet_name}")
                    return []
        except Exception as e:
            self.logger.error(f"Ошибка при определении колонки с наименованиями: {e}")
            return []
            
        self.logger.info(f"Используем колонку '{item_column_name}' для наименований на листе {self.sheet_name}")
            
        # Проходим по всем строкам, пропуская указанное количество строк заголовка
        for idx, row in enumerate(self.df.iterrows()):
            idx, row = row  # Распаковываем кортеж (индекс, Series)
            total_rows += 1
            
            # Пропускаем первые N строк (заголовки таблицы)
            if idx < self._num_header_rows:
                headers_skipped += 1
                self.logger.debug(f"Пропускаем строку {idx} (часть заголовка)")
                continue
            
            # Проверяем наличие колонки в row
            if item_column_name not in row:
                self.logger.warning(f"Колонка '{item_column_name}' отсутствует в строке {idx}")
                empty_skipped += 1
                continue
                
            # Пропускаем пустые ячейки
            if pd.isna(row[item_column_name]):
                empty_skipped += 1
                continue
                
            item_text = str(row[item_column_name]).strip()
            
            # Пропускаем пустые строки и одиночные символы
            if not item_text or len(item_text) <= 1:
                empty_skipped += 1
                continue
                
            # Пропускаем, если это число или короткое число
            if item_text.isdigit() and len(item_text) <= 3:
                empty_skipped += 1
                continue
            
            # Проверяем по шаблонам служебных строк
            skip_item = False
            for i, pattern in enumerate(self.SKIP_PATTERNS_COMPILED):
                if pattern.search(item_text):
                    skip_item = True
                    patterns_skipped += 1
                    self.logger.info(f"Пропускаем служебную строку [{idx}]: '{item_text}' (соответствует шаблону {i+1}: {self.SKIP_PATTERNS[i]})")
                    break
                    
            if skip_item:
                continue
                
            # Добавляем элемент в список
            items_collected.append((idx, item_text))
        
        # Обновляем счетчики для логирования
        self.skipped_rows = headers_skipped + empty_skipped
        self.skipped_service_rows = patterns_skipped
        
        self.logger.info(f"Всего строк на листе: {total_rows}")
        self.logger.info(f"Пропущено строк заголовка: {headers_skipped}")
        self.logger.info(f"Пропущено пустых строк: {empty_skipped}")
        self.logger.info(f"Пропущено служебных строк: {patterns_skipped}")
        self.logger.info(f"Собрано элементов: {len(items_collected)}")
        
        return items_collected
    
    def _update_excel_with_codes(self):
        """
        Обновляет коды ОКПД в исходном Excel-файле, сохраняя форматирование
        
        Returns:
            bool: True если успешно, False в случае ошибки
        """
        if not self.results_to_update:
            self.logger.warning("Нет данных для обновления Excel файла")
            return False
            
        if not self.input_path or not os.path.exists(self.input_path):
            self.logger.error("Не найден исходный Excel файл")
            return False
            
        try:
            # Если есть открытый workbook, используем его
            if not self.workbook:
                self.workbook = openpyxl.load_workbook(self.input_path)
                
            # Используем указанный лист, активный лист или первый лист
            if self.sheet_name and self.sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[self.sheet_name]
            else:
                sheet = self.workbook.active
                self.sheet_name = sheet.title
                
            self.logger.info(f"Обновление данных в листе '{sheet.title}'")
            
            # Счетчик обновлений
            updated = 0
            
            # Создаем словарь для поиска текстов по содержимому листа
            cell_content_map = {}
            items_found = set()
            
            # Список всех листов для поиска
            sheets_to_search = [sheet]
            
            # Если не нашли текст на текущем листе, попробуем поискать на других листах
            if len(self.results_to_update) > 0 and self.workbook.sheetnames:
                for sheet_name in self.workbook.sheetnames:
                    if sheet_name != sheet.title:
                        other_sheet = self.workbook[sheet_name]
                        sheets_to_search.append(other_sheet)
                        self.logger.info(f"Добавлен лист '{sheet_name}' для поиска элементов")
            
            # Выводим список элементов для отладки
            if len(self.results_to_update) < 10:
                elements_to_find = [data['item'] for _, data in self.results_to_update.items()]
                self.logger.info(f"Ищем элементы: {elements_to_find}")
            else:
                self.logger.info(f"Ищем {len(self.results_to_update)} элементов")
            
            # Проверяем каждый лист для построения карты
            for current_sheet in sheets_to_search:
                self.logger.info(f"Сканирование листа '{current_sheet.title}' для поиска элементов")
                
                for row in range(1, current_sheet.max_row + 1):
                    # Ищем ячейку с названием элемента
                    name_col = self.name_column_index
                    cell = current_sheet.cell(row=row, column=name_col)
                    
                    if cell.value:
                        # Нормализуем текст ячейки для лучшего сравнения
                        cell_text = str(cell.value)
                        normalized_text = self._normalize_cell_text(cell_text)
                        
                        if normalized_text:
                            # Сохраняем строку и лист
                            cell_content_map[normalized_text] = (current_sheet, row)
                            
                            # Добавляем также версию без пробелов
                            no_spaces = normalized_text.replace(" ", "")
                            if no_spaces != normalized_text:
                                cell_content_map[no_spaces] = (current_sheet, row)
                            
                            # Проверяем, нашли ли мы какой-то из искомых элементов
                            for _, data in self.results_to_update.items():
                                item_text = data['item']
                                item_norm = self._normalize_cell_text(item_text)
                                
                                if (item_norm == normalized_text or 
                                    item_norm.replace(" ", "") == no_spaces or
                                    normalized_text in item_norm or 
                                    item_norm in normalized_text):
                                    items_found.add(item_text)
            
            # Отладочная информация о найденных элементах
            found_percent = len(items_found) / len(self.results_to_update) * 100 if self.results_to_update else 0
            self.logger.info(f"Построена карта содержимого листов: найдено {len(items_found)} из {len(self.results_to_update)} элементов ({found_percent:.1f}%)")
            
            # Если не найдены некоторые элементы, выведем их для отладки
            if len(items_found) < len(self.results_to_update) and len(self.results_to_update) - len(items_found) < 10:
                missing = [data['item'] for _, data in self.results_to_update.items() if data['item'] not in items_found]
                self.logger.warning(f"Не найдены элементы: {missing}")
            elif len(items_found) < len(self.results_to_update):
                self.logger.warning(f"Не найдено {len(self.results_to_update) - len(items_found)} элементов")
            
            for row_idx, data in self.results_to_update.items():
                code = data['code']
                column_name = data['column_name']
                item_text = data['item']
                
                # Нормализуем текст элемента для лучшего сравнения
                item_normalized = self._normalize_cell_text(item_text)
                item_no_spaces = item_normalized.replace(" ", "")
                
                # Определяем номер колонки для кода ОКПД
                if isinstance(column_name, int):
                    # Если индекс pandas (с 0), преобразуем в индекс openpyxl (с 1)
                    col_idx = column_name + 1
                else:
                    # Если это название колонки, найдем соответствующий индекс
                    col_idx = self.code_column_index
                
                if not col_idx:
                    self.logger.warning(f"Не удалось определить колонку для кода ОКПД")
                    continue
                
                # Номер строки в Excel (строка в pandas + header_rows + 1 для учета индексации с 0)
                excel_row = int(row_idx) + 1
                target_sheet = sheet  # По умолчанию текущий лист
                
                # Проверяем, найден ли элемент в карте содержимого
                found = False
                
                # Пробуем точное совпадение
                if item_normalized in cell_content_map:
                    target_sheet, excel_row = cell_content_map[item_normalized]
                    self.logger.info(f"Найдено точное соответствие для '{item_text}' в листе '{target_sheet.title}', строка {excel_row}")
                    found = True
                # Пробуем версию без пробелов
                elif item_no_spaces in cell_content_map:
                    target_sheet, excel_row = cell_content_map[item_no_spaces]
                    self.logger.info(f"Найдено соответствие без пробелов для '{item_text}' в листе '{target_sheet.title}', строка {excel_row}")
                    found = True
                else:
                    # Если точного совпадения нет, попробуем найти по частичному совпадению
                    best_match = None
                    best_ratio = 0.8  # Минимальный порог сходства (80%)
                    best_key = None
                    
                    # Поиск методом частичного совпадения
                    for key, (s, row) in cell_content_map.items():
                        # Пропускаем короткие строки
                        if len(key) < 5:
                            continue
                            
                        # Проверяем, содержит ли текст ячейки наш элемент
                        if item_normalized in key or key in item_normalized:
                            ratio = len(min(item_normalized, key, key=len)) / len(max(item_normalized, key, key=len))
                            if ratio > best_ratio:
                                best_ratio = ratio
                                best_match = (s, row)
                                best_key = key
                    
                    if best_match:
                        target_sheet, excel_row = best_match
                        cell_value = target_sheet.cell(row=excel_row, column=self.name_column_index).value
                        self.logger.info(f"Найдено частичное соответствие для '{item_text}' в листе '{target_sheet.title}', строка {excel_row}: '{cell_value}' (совпадение {best_ratio:.2f})")
                        found = True
                
                # Если не найдено совпадение, пропускаем элемент
                if not found:
                    self.logger.warning(f"Не удалось найти строку с текстом '{item_text}' ни в одном листе, пропускаем")
                    continue
                
                # Записываем код в ячейку
                try:
                    # Если лист отличается от текущего, запоминаем это
                    if target_sheet.title != sheet.title:
                        self.logger.info(f"Переключаемся на лист '{target_sheet.title}' для обновления ячейки")
                    
                    cell = target_sheet.cell(row=excel_row, column=col_idx)
                    old_value = cell.value
                    
                    # Проверяем, содержит ли ячейка формулу
                    if old_value and isinstance(old_value, str) and old_value.startswith('='):
                        self.logger.warning(f"Ячейка ({target_sheet.title}:{excel_row}, {col_idx}) содержит формулу, пропускаем: {old_value}")
                        continue
                    
                    cell.value = code
                    updated += 1
                    self.logger.debug(f"Обновлена ячейка {target_sheet.title}:({excel_row}, {col_idx}): '{old_value}' -> '{code}'")
                except Exception as e:
                    self.logger.warning(f"Ошибка при обновлении ячейки {target_sheet.title}:({excel_row}, {col_idx}): {e}")
            
            # Сохраняем изменения
            self.workbook.save(self.input_path)
            self.logger.info(f"Файл Excel обновлен: {updated} кодов ОКПД добавлено")
            
            return True
            
        except Exception as e:
            self.logger.exception(f"Ошибка при обновлении Excel файла: {e}")
            return False
            
    def _normalize_cell_text(self, text):
        """Нормализует текст ячейки для сравнения"""
        if not text:
            return ""
            
        # Преобразуем в строку
        text = str(text)
        
        # Заменяем неразрывные пробелы на обычные
        text = text.replace('\xa0', ' ')
        
        # Убираем лишние пробелы
        text = " ".join(text.split())
        
        # Удаляем непечатаемые символы
        text = ''.join(c for c in text if c.isprintable())
        
        # Игнорируем специальные символы, которые могут различаться
        text = text.replace('-', ' ').replace('_', ' ').replace('.', ' ').replace(',', ' ')
        text = " ".join(text.split())
        
        return text.strip().lower()  # Приводим к нижнему регистру для регистронезависимого сравнения 

    def _adjust_column_width(self):
        """
        Увеличивает ширину колонки с кодами ОКПД в 3 раза на всех листах
        """
        try:
            if not self.workbook:
                self.logger.warning("Рабочая книга не открыта, невозможно изменить ширину колонки")
                return False
            
            self.logger.info("Увеличение ширины колонки с кодами ОКПД на всех листах")
            
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                
                # Находим колонку с кодами ОКПД на текущем листе
                # Временно сохраняем текущие значения
                current_sheet_name = self.sheet_name
                current_code_column_index = self.code_column_index
                
                # Устанавливаем текущий лист для поиска колонок
                self.sheet_name = sheet_name
                success = self._find_columns_in_excel()
                
                if success and self.code_column_index:
                    # Получаем букву колонки из ее индекса
                    col_letter = get_column_letter(self.code_column_index)
                    
                    # Получаем текущую ширину колонки
                    current_width = sheet.column_dimensions[col_letter].width
                    
                    # Если ширина не задана, используем значение по умолчанию (примерно 8.43)
                    if not current_width or current_width < 0:
                        current_width = 8.43
                    
                    # Увеличиваем ширину в 3 раза
                    new_width = current_width * 3
                    
                    # Устанавливаем новую ширину
                    sheet.column_dimensions[col_letter].width = new_width
                    
                    self.logger.info(f"Лист '{sheet_name}': ширина колонки {col_letter} изменена с {current_width} на {new_width}")
                else:
                    self.logger.warning(f"Не удалось найти колонку с кодами ОКПД на листе '{sheet_name}'")
                
                # Восстанавливаем предыдущие значения
                self.sheet_name = current_sheet_name
                self.code_column_index = current_code_column_index
            
            self.logger.info("Ширина колонок с кодами ОКПД успешно увеличена на всех листах")
            return True
            
        except Exception as e:
            self.logger.exception(f"Ошибка при изменении ширины колонки: {e}")
            return False 