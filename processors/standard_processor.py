"""
Обработчик для стандартного формата Excel файлов
"""

import pandas as pd
import os
import time
import openpyxl
import shutil
from main import Processor, group_similar
from src.morphology import normalize_term
from src.okpd_fetch import fetch_okpd2_batch
from .base_processor import BaseProcessor

class StandardProcessor(BaseProcessor):
    """Процессор для стандартного формата файлов с колонкой 'Наименование'"""
    
    def __init__(self, input_file=None, checkpoint_name="checkpoint.xlsx", save_interval=10, progress=None):
        super().__init__(input_file, checkpoint_name, save_interval, progress)
        
        # Данные для обновления Excel файла
        self.results_to_update = {}
        
        # Данные о файле
        self.workbook = None
        self.sheet_name = None
        self.name_column_index = None
        self.code_column_index = None
        
        # Создаем резервную копию входного файла сразу
        if input_file and input_file.name:
            self._create_backup_file()
    
    def _create_backup_file(self):
        """Создает резервную копию исходного файла"""
        try:
            original_path = self.input_path
            if not original_path or not os.path.exists(original_path):
                return
                
            # Формируем имя для бэкапа
            base_name, ext = os.path.splitext(original_path)
            backup_path = f"{base_name}_original{ext}"
            
            # Создаем копию только если её ещё нет
            if not os.path.exists(backup_path):
                shutil.copy2(original_path, backup_path)
                self.logger.info(f"Создана резервная копия исходного файла: {backup_path}")
                return backup_path
        except Exception as e:
            self.logger.warning(f"Не удалось создать резервную копию: {e}")
        return None
    
    def _find_columns_in_excel(self):
        """
        Находит колонки с наименованием и кодами ОКПД в файле Excel
        непосредственно используя openpyxl
        """
        if not self.input_path or not os.path.exists(self.input_path):
            self.logger.error("Не указан путь к входному файлу")
            return False
            
        try:
            # Открываем Excel файл
            self.workbook = openpyxl.load_workbook(self.input_path)
            self.logger.info(f"Excel файл открыт: {self.input_path}")
            
            # Определяем активный лист (или первый, если активный не задан)
            if self.workbook.active:
                sheet = self.workbook.active
                self.sheet_name = sheet.title
            else:
                sheet = self.workbook.worksheets[0]
                self.sheet_name = sheet.title
            
            self.logger.info(f"Анализируем лист: {self.sheet_name}")
            
            # Ищем в первых нескольких строках заголовки колонок
            name_column = None
            code_column = None
            
            # Проверяем первые 15 строк (обычно заголовки там)
            max_check_row = min(15, sheet.max_row)
            
            for row in range(1, max_check_row + 1):
                for col in range(1, min(20, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    
                    if not cell_value:
                        continue
                        
                    cell_text = str(cell_value).strip().lower()
                    
                    # Ищем колонку с наименованием
                    if cell_text == "наименование" or "наименов" in cell_text:
                        name_column = col
                        self.logger.info(f"Найдена колонка 'Наименование': строка {row}, колонка {col}")
                    
                    # Ищем колонку с кодом ОКПД
                    if "окпд" in cell_text or "код окп" in cell_text:
                        code_column = col
                        self.logger.info(f"Найдена колонка с кодом ОКПД: строка {row}, колонка {col}")
                    
                    # Если нашли обе колонки, завершаем поиск
                    if name_column and code_column:
                        self.name_column_index = name_column
                        self.code_column_index = code_column
                        return True
            
            # Если не нашли колонку кода, но нашли наименование
            if name_column and not code_column:
                self.name_column_index = name_column
                # Ищем подходящую колонку для кодов
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=1, column=col).value
                    if cell_value and 'код' in str(cell_value).lower():
                        code_column = col
                        self.code_column_index = col
                        self.logger.info(f"Найдена колонка для кодов: {col}")
                        return True
                        
                # Если не нашли подходящую, создаем новую колонку после наименования
                code_column = name_column + 1
                self.code_column_index = code_column
                self.logger.info(f"Будем использовать колонку {code_column} для кодов ОКПД")
                
                # Добавляем заголовок в ячейку для кодов ОКПД
                if sheet.cell(row=1, column=name_column).value:
                    sheet.cell(row=1, column=code_column).value = "ОКПД код"
                    self.logger.info(f"Добавлен заголовок 'ОКПД код' в строку 1, колонку {code_column}")
                
                return True
            
            # Не нашли нужные колонки
            if not name_column:
                self.logger.error("Не удалось найти колонку 'Наименование'")
                
                # Поиск по содержимому - ищем колонку с наибольшим количеством текста
                text_counts = {}
                for col in range(1, min(10, sheet.max_column + 1)):
                    text_count = 0
                    for row in range(1, min(20, sheet.max_row + 1)):
                        cell_value = sheet.cell(row=row, column=col).value
                        if cell_value and isinstance(cell_value, str) and len(str(cell_value).strip()) > 5:
                            text_count += 1
                    text_counts[col] = text_count
                
                if text_counts:
                    # Находим колонку с наибольшим количеством текста
                    name_column = max(text_counts.items(), key=lambda x: x[1])[0]
                    if text_counts[name_column] > 3:  # Если есть хотя бы несколько текстовых ячеек
                        self.name_column_index = name_column
                        self.logger.info(f"По содержимому определена колонка наименований: {name_column}")
                        
                        # Для кодов используем следующую колонку
                        code_column = name_column + 1
                        self.code_column_index = code_column
                        self.logger.info(f"Для кодов ОКПД будем использовать колонку {code_column}")
                        return True
            
            return False
            
        except Exception as e:
            self.logger.exception(f"Ошибка при анализе Excel файла: {e}")
            return False
    
    def _process_file(self):
        """Обработка файла стандартного формата"""
        try:
            start_time = time.time()
            self.logger.info(f"Начало обработки стандартного файла: {self.input_path}")
            
            # Анализируем файл с помощью openpyxl
            if not self._find_columns_in_excel():
                self.logger.error("Не удалось найти необходимые колонки в файле")
                return False
            
            # Чтение файла pandas для обработки данных
            read_start = time.time()
            try:
                self.df = pd.read_excel(self.input_path)
                read_time = time.time() - read_start
                self.logger.info(f"Файл прочитан для анализа за {read_time:.1f} сек. Обнаружено {self.df.shape[0]} строк, {self.df.shape[1]} столбцов")
            except Exception as e:
                self.logger.exception(f"Ошибка при чтении файла pandas: {e}")
                return False
                
            # Определяем колонки в pandas DataFrame
            name_col = None
            
            # Если нашли колонку наименования через openpyxl
            if self.name_column_index:
                pandas_col_idx = self.name_column_index - 1  # Индексы в pandas начинаются с 0
                
                # Проверяем, есть ли такой индекс в DataFrame
                if pandas_col_idx < len(self.df.columns):
                    name_col = self.df.columns[pandas_col_idx]
                    self.logger.info(f"Используем колонку '{name_col}' (индекс {pandas_col_idx}) для наименований")
            
            # Если не нашли через индекс, ищем по имени
            if not name_col:
                if 'Наименование' in self.df.columns:
                    name_col = 'Наименование'
                    self.logger.info(f"Используем колонку 'Наименование'")
                else:
                    # Ищем колонку с наименованием по подстроке
                    name_cols = [col for col in self.df.columns if 'наименов' in str(col).lower()]
                    if name_cols:
                        name_col = name_cols[0]
                        self.logger.info(f"Используем колонку '{name_col}' для наименований")
                    else:
                        self.logger.error("Не удалось найти колонку с наименованиями в DataFrame")
                        return False
            
            # Получение терминов
            originals = self.df[name_col].dropna().tolist()
            self.logger.info(f"Обнаружено {len(originals)} элементов для обработки")
            
            if not originals:
                self.logger.error(f"Колонка '{name_col}' не содержит данных для обработки")
                return False
            
            # Группировка терминов
            group_start = time.time()
            groups = group_similar(originals)
            group_time = time.time() - group_start
            self.logger.info(f"Сгруппировано в {len(groups)} групп за {group_time:.1f} сек")
            
            # Инициализация модели
            if not self.init_model():
                self.logger.error("Не удалось инициализировать модель")
                return False
                
            # Инициализация для хранения результатов
            self.results_to_update = {}
            name_to_idx = {}
            
            # Создаем отображение наименований на индексы строк
            for idx, row in self.df.iterrows():
                if pd.notna(row[name_col]):
                    item_text = str(row[name_col]).strip()
                    if item_text:
                        name_to_idx[item_text] = idx
            
            # Инициализация индикатора прогресса
            total = len(groups)
            processed_items = 0
            success_items = 0
            error_items = 0
            
            # Безопасно работаем с объектом прогресса
            if self.progress is not None:
                self.progress(0.0, desc="Инициализация обработки...", total=total)
                progress_iter = self.progress.tqdm(groups, desc="Обработка групп")
            else:
                progress_iter = groups
            
            # Обработка каждой группы
            process_start = time.time()
            for idx, grp in enumerate(progress_iter, start=1):
                if self.stop_event.is_set():
                    self.logger.info("Обработка остановлена пользователем")
                    break
                    
                if not grp:
                    continue
                    
                rep = grp[0]
                processed_items += len(grp)
                
                try:
                    # Обработка термина
                    normalized = normalize_term(rep)
                    self.logger.info(f"Обработка группы {idx}/{total}: {normalized}")
                    
                    # Получение упрощенного термина
                    prompt = [{"role": "user", "content": normalized}]
                    simplified = self.model.generate(prompt)['content']
                    self.logger.info(f"Упрощено до: {simplified}")
                    
                    # Получение кодов ОКПД
                    okpd_data = fetch_okpd2_batch([simplified])
                    entries = okpd_data.get(simplified, [])
                    
                    # Выбор подходящего кода
                    code, name, comment = Processor._decide(None, entries, rep, simplified)
                    self.logger.info(f"Выбран код: {code} - {name}")
                    
                    # Добавляем результаты в словарь для обновления Excel
                    for item in grp:
                        if item in name_to_idx:
                            row_idx = name_to_idx[item]
                            self.results_to_update[row_idx] = {
                                'code': code,
                                'name': name,
                                'comment': comment,
                                'item': item
                            }
                            success_items += 1
                    
                    # Обновление статуса прогресса
                    if self.progress is not None:
                        percent = (idx / total) * 100
                        remaining = total - idx
                        self.progress(
                            idx / total, 
                            desc=f"Обработано {idx}/{total} групп ({percent:.1f}%), осталось {remaining}"
                        )
                    
                    # Сохранение промежуточных результатов
                    if idx % int(self.save_interval) == 0:
                        self._update_excel_with_codes()
                        self.logger.info(f"Сохранен промежуточный результат: группа {idx}/{total}, обработано {processed_items} элементов")
                        
                except Exception as e:
                    self.logger.exception(f"Ошибка при обработке группы {idx}: {e}")
                    error_items += len(grp)
            
            process_time = time.time() - process_start
            self.logger.info(f"Обработка закончена за {process_time:.1f} сек")
            self.logger.info(f"Итого: обработано {success_items} из {processed_items} элементов, ошибок: {error_items}")
            
            # Обновляем Excel файл с кодами
            if self.results_to_update:
                result = self._update_excel_with_codes()
                
                if result:
                    self.logger.info(f"Результаты успешно записаны в исходный файл с сохранением форматирования")
                    
                    # Копируем исходный файл в output_path для интерфейса
                    try:
                        shutil.copy2(self.input_path, self.output_path)
                        self.logger.info(f"Копия результата сохранена в {self.output_path}")
                    except Exception as e:
                        self.logger.warning(f"Не удалось создать копию результата: {e}")
                
                    # Общее время обработки
                    total_time = time.time() - start_time
                    self.logger.info(f"Общее время обработки: {total_time:.1f} сек")
                    
                    return True
                else:
                    self.logger.error("Ошибка при записи результатов в исходный файл")
                    # Попробуем сохранить стандартным методом
                    self._save_results_to_new_file()
                    return False
            else:
                self.logger.warning("Нет данных для обновления в Excel")
                return False
                    
        except Exception as e:
            self.logger.exception(f"Критическая ошибка при обработке: {e}")
            return False
    
    def _update_excel_with_codes(self):
        """
        Обновляет коды ОКПД в исходном Excel-файле, сохраняя форматирование
        
        Returns:
            bool: True если успешно, False в случае ошибки
        """
        if not self.results_to_update:
            self.logger.warning("Нет данных для обновления Excel файла")
            return False
            
        if not self.input_path or not os.path.exists(self.input_path):
            self.logger.error("Не найден исходный Excel файл")
            return False
            
        try:
            # Если есть открытый workbook, используем его
            if not self.workbook:
                self.workbook = openpyxl.load_workbook(self.input_path)
                
            # Используем активный лист или первый лист
            if self.sheet_name:
                if self.sheet_name in self.workbook.sheetnames:
                    sheet = self.workbook[self.sheet_name]
                else:
                    sheet = self.workbook.active
            else:
                sheet = self.workbook.active
                
            self.logger.info(f"Обновление данных в листе '{sheet.title}'")
            
            # Определяем колонки для кодов
            if not self.code_column_index:
                self.logger.warning("Не определена колонка для кодов ОКПД")
                return False
                
            code_col = self.code_column_index
            name_col = self.code_column_index + 1  # Колонка для названия кода
            comment_col = self.code_column_index + 2  # Колонка для комментария
            
            # Проверяем, есть ли заголовки у колонок
            first_row = sheet[1]
            if not sheet.cell(row=1, column=code_col).value:
                sheet.cell(row=1, column=code_col).value = "ОКПД код"
            if not sheet.cell(row=1, column=name_col).value:
                sheet.cell(row=1, column=name_col).value = "Название кода"
            if not sheet.cell(row=1, column=comment_col).value:
                sheet.cell(row=1, column=comment_col).value = "Комментарий"
            
            # Счетчик обновлений
            updated = 0
            
            for row_idx, data in self.results_to_update.items():
                code = data.get('code', '')
                name = data.get('name', '')
                comment = data.get('comment', '')
                
                # Индекс строки в Excel (строка в pandas + 1, т.к. openpyxl считает с 1)
                excel_row = int(row_idx) + 1
                
                # Записываем значения в ячейки
                try:
                    sheet.cell(row=excel_row, column=code_col).value = code
                    sheet.cell(row=excel_row, column=name_col).value = name
                    sheet.cell(row=excel_row, column=comment_col).value = comment
                    updated += 1
                except Exception as e:
                    self.logger.warning(f"Ошибка при обновлении ячеек в строке {excel_row}: {e}")
            
            # Сохраняем изменения
            self.workbook.save(self.input_path)
            self.logger.info(f"Файл Excel обновлен: {updated} элементов получили коды ОКПД")
            
            return True
            
        except Exception as e:
            self.logger.exception(f"Ошибка при обновлении Excel файла: {e}")
            return False
    
    def _save_results_to_new_file(self):
        """Сохраняет результаты в новый Excel файл (резервный метод)"""
        try:
            self.logger.info(f"Создание нового файла с результатами: {self.output_path}")
            
            # Проверяем, есть ли данные для сохранения
            if not hasattr(self, 'df') or self.df is None:
                self.logger.error("Нет данных для сохранения")
                return False
            
            # Создаем новый DataFrame с результатами
            result_df = self.df.copy()
            
            # Добавляем колонки для результатов, если их нет
            for col in ['ОКПД код', 'Название кода', 'Комментарий']:
                if col not in result_df.columns:
                    result_df[col] = ''
            
            # Заполняем результаты
            for row_idx, data in self.results_to_update.items():
                result_df.loc[row_idx, 'ОКПД код'] = data.get('code', '')
                result_df.loc[row_idx, 'Название кода'] = data.get('name', '')
                result_df.loc[row_idx, 'Комментарий'] = data.get('comment', '')
            
            # Сохраняем в новый файл
            result_df.to_excel(self.output_path, index=False)
            self.logger.info(f"Результаты сохранены в новый файл: {self.output_path}")
            
            return True
        except Exception as e:
            self.logger.exception(f"Ошибка при создании нового файла с результатами: {e}")
            return False 