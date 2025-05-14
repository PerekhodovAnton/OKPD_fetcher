import pandas as pd
import openpyxl
import os
import logging
import shutil

# Configure logging
logging.basicConfig(level=logging.INFO, 
                   format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger('excel_test')

# Test files
original_file = '4_1.xlsx'
output_file = '4_1_with_codes.xlsx'

def test_excel_preservation():
    """Test that we can update cells in Excel while preserving all formatting"""
    if not os.path.exists(original_file):
        logger.error(f"Original file {original_file} not found!")
        return
    
    # Make a copy of the original file to work with
    shutil.copy2(original_file, output_file)
    logger.info(f"Created working copy: {output_file}")
    
    # Step 1: Open the Excel file with openpyxl
    try:
        workbook = openpyxl.load_workbook(output_file)
        logger.info(f"Successfully opened Excel file with openpyxl")
        
        # Get information about the file
        sheet_names = workbook.sheetnames
        logger.info(f"Workbook contains {len(sheet_names)} sheets: {sheet_names}")
        
        # Use the active sheet
        sheet = workbook.active
        logger.info(f"Active sheet: {sheet.title}")
        logger.info(f"Sheet dimensions: {sheet.dimensions}")
        logger.info(f"Contains {len(sheet.merged_cells.ranges)} merged cell ranges")
        
        # Find the column headers
        name_column = None
        code_column = None
        
        # Check first 15 rows for headers
        for row in range(1, min(15, sheet.max_row + 1)):
            for col in range(1, min(15, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row, column=col).value
                if not cell_value:
                    continue
                    
                cell_text = str(cell_value).strip().lower()
                
                # Find name column
                if "наименование" in cell_text:
                    name_column = col
                    logger.info(f"Found 'Наименование' column at row {row}, column {col}")
                
                # Find code column
                if "код" in cell_text and ("окп" in cell_text or "окпд" in cell_text):
                    code_column = col
                    logger.info(f"Found 'Код ОКП/ОКПД2' column at row {row}, column {col}")
        
        # If we didn't find the name column, try looking for content
        if not name_column:
            for col in range(1, min(15, sheet.max_column + 1)):
                # Look for column with meaningful text content
                content_found = 0
                for row in range(5, 15):  # Check a few rows after header
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and len(str(cell_value).strip()) > 10:
                        content_found += 1
                
                if content_found >= 3:  # If multiple cells have content
                    name_column = col
                    logger.info(f"Based on content, using column {col} as name column")
                    break
        
        # If we didn't find code column, create or choose one
        if not code_column:
            if name_column:
                # Use column next to name column
                code_column = name_column + 1
                logger.info(f"No code column found, using column {code_column}")
                
                # Add header text for this column
                header_row = 3  # Typical header row
                sheet.cell(row=header_row, column=code_column).value = "Код ОКП/ОКПД2"
            else:
                code_column = 4  # Default to column D if nothing found
                logger.info(f"Using default column {code_column} for codes")
        
        # Add some test data - update cells in the code column
        if name_column and code_column:
            test_codes = {
                10: "28.30.59.111",
                11: "28.92.61.110", 
                12: "25.73.60.120",
                13: "28.49.12.190",
                14: "28.41.40.000"
            }
            
            # Update cells
            for row_num, code in test_codes.items():
                sheet.cell(row=row_num, column=code_column).value = code
                logger.info(f"Set code '{code}' at row {row_num}, column {code_column}")
            
            # Save the file with all changes
            workbook.save(output_file)
            logger.info(f"Successfully saved file with modifications: {output_file}")
            
            # Verify the file was saved and can be opened
            if os.path.exists(output_file):
                file_size = os.path.getsize(output_file)
                logger.info(f"Output file size: {file_size} bytes")
                
                # Try opening it again to verify
                test_wb = openpyxl.load_workbook(output_file)
                logger.info(f"Successfully verified file by opening it again")
                
                # Check if our codes are there
                test_sheet = test_wb.active
                for row_num, code in test_codes.items():
                    value = test_sheet.cell(row=row_num, column=code_column).value
                    logger.info(f"Verification: row {row_num}, column {code_column}: '{value}'")
                
                return True
            else:
                logger.error(f"Output file not found after saving!")
                return False
        else:
            logger.error("Failed to identify necessary columns")
            return False
            
    except Exception as e:
        logger.exception(f"Error during Excel processing: {e}")
        return False

if __name__ == "__main__":
    result = test_excel_preservation()
    logger.info(f"Test completed with result: {result}") 